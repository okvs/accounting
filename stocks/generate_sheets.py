# -*- coding: utf-8 -*-
import glob
import os
import re
import traceback

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

STOCKS_DIR = r'C:\Workspace\accounting\stocks'
OUTPUT_PATH = os.path.join(STOCKS_DIR, 'stocks_output.xlsx')
SKIP_PATTERNS = ['예시']
SPC_COL = '수익증권/spc명'
LOG_COL = 'log'


def normalize_sheet_name(name):
    return re.sub(r'\s+', '', name)


def is_skip_sheet(name):
    for pat in SKIP_PATTERNS:
        if pat in name:
            return True
    return False


def find_company_name(ws):
    for col in range(8, 12):  # H=8 ~ K=11
        v = ws.cell(1, col).value
        if v is not None and str(v).strip():
            return str(v).strip()
    return ''


def find_header_row(ws):
    last_match = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(r, c).value
            if v and re.match(r'계정과목', str(v)):
                all_vals = [ws.cell(r, col).value for col in range(1, ws.max_column + 1)]
                non_none = [x for x in all_vals if x is not None]
                if len(non_none) >= 3:
                    last_match = r
    return last_match


def extract_detail_data(ws, header_row):
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None:
            h = str(v).replace(' ▼', '').strip()
            headers.append((c, h))

    if not headers:
        return pd.DataFrame()

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col_idx, col_name in headers:
            v = ws.cell(r, col_idx).value
            if v is not None and str(v).strip() and str(v).strip() != '-':
                has_data = True
            row_data[col_name] = v
        if has_data:
            all_text = ' '.join(str(v) for v in row_data.values() if v is not None)
            if re.search(r'\d+\s*페이지', all_text):
                continue
            rows.append(row_data)

    return pd.DataFrame(rows, columns=[h for _, h in headers])


def get_category_from_a1(ws):
    v = ws.cell(1, 1).value
    if v and '결산명세_' in str(v):
        return str(v).split('결산명세_', 1)[1].strip()
    return None


def auto_adjust_column_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                char_len = 0
                for ch in val_str:
                    char_len += 2.1 if ord(ch) > 127 else 1
                max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def format_numbers_as_text(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float) and cell.value == int(cell.value):
                cell.value = int(cell.value)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'


def main():
    logs = []

    xlsx_files = sorted(glob.glob(os.path.join(STOCKS_DIR, '*.xlsx')))
    xlsx_files = [
        f for f in xlsx_files
        if os.path.basename(f) != os.path.basename(OUTPUT_PATH)
        and not os.path.basename(f).startswith('~$')
    ]

    if not xlsx_files:
        print('처리할 엑셀 파일이 없습니다.')
        return

    print(f'대상 파일 {len(xlsx_files)}개:')
    for f in xlsx_files:
        print(f'  - {os.path.basename(f)}')

    combined_by_category = {}
    source_counts = {}

    for fpath in xlsx_files:
        fname = os.path.basename(fpath)
        print(f'\n=== {fname} ===')

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            msg = f'[ERROR] {fname}: 파일 열기 실패 - {e}'
            print(msg)
            logs.append(msg)
            logs.append(traceback.format_exc())
            continue

        for sheet_name in wb.sheetnames:
            if is_skip_sheet(sheet_name):
                print(f'  [SKIP] {sheet_name}: 예시 시트')
                continue

            try:
                ws = wb[sheet_name]

                category = get_category_from_a1(ws)
                if not category:
                    category = normalize_sheet_name(sheet_name)

                company_name = find_company_name(ws)
                if not company_name:
                    msg = f'[WARN] {fname}/{sheet_name}: 회사명(H~K열)을 찾을 수 없음'
                    print(f'  {msg}')
                    logs.append(msg)

                header_row = find_header_row(ws)

                if header_row is None:
                    msg = f'[WARN] {fname}/{sheet_name}: 계정과목 헤더 행을 찾을 수 없음'
                    print(f'  {msg}')
                    logs.append(msg)
                    continue

                df = extract_detail_data(ws, header_row)

                if df.empty:
                    msg = f'[WARN] {fname}/{sheet_name}: 데이터 행 없음 (헤더 행={header_row})'
                    print(f'  {msg}')
                    logs.append(msg)
                    continue

                df[SPC_COL] = company_name
                norm_cat = normalize_sheet_name(category)

                if norm_cat not in combined_by_category:
                    combined_by_category[norm_cat] = []
                    source_counts[norm_cat] = []
                combined_by_category[norm_cat].append(df)
                source_counts[norm_cat].append((fname, sheet_name, len(df)))
                print(f'  [OK] {sheet_name} -> {norm_cat}: {len(df)}행, 회사명={company_name}')

            except Exception as e:
                msg = f'[ERROR] {fname}/{sheet_name}: {e}'
                print(f'  {msg}')
                logs.append(msg)
                logs.append(traceback.format_exc())

        wb.close()

    if not combined_by_category:
        msg = '추출된 데이터가 없습니다.'
        print(f'\n{msg}')
        logs.append(msg)
        return

    merged_sheets = {}
    for cat, dfs in combined_by_category.items():
        merged_sheets[cat] = pd.concat(dfs, ignore_index=True)

    print('\n=== 데이터 건수 검증 ===')
    for cat, counts in source_counts.items():
        source_total = sum(c for _, _, c in counts)
        output_total = len(merged_sheets[cat])
        status = 'OK' if source_total == output_total else 'MISMATCH'
        detail = ' + '.join(f'{fn}/{sn}({c}행)' for fn, sn, c in counts)
        line = f'  [{status}] {cat}: 원본 {detail} = {source_total}행 → 출력 {output_total}행'
        print(line)
        if status == 'MISMATCH':
            msg = f'[ERROR] {cat}: 건수 불일치 - 원본 {source_total}행 vs 출력 {output_total}행'
            logs.append(msg)

    all_dfs = []
    for cat, df in merged_sheets.items():
        df_copy = df.copy()
        df_copy.insert(0, '시트명', cat)
        all_dfs.append(df_copy)
    combined_all = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

    log_status = '정상' if not logs else '\n'.join(logs)

    try:
        open(OUTPUT_PATH, 'a').close()
    except PermissionError:
        print(f'\n[ERROR] 출력 파일이 다른 프로그램에서 열려 있습니다: {OUTPUT_PATH}')
        print('Excel을 닫고 다시 실행해주세요.')
        return

    with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
        if not combined_all.empty:
            combined_all[LOG_COL] = log_status
            combined_all.to_excel(writer, sheet_name='전체', index=False)

        for cat, df in merged_sheets.items():
            safe_name = cat[:31]
            cat_logs = [l for l in logs if cat in l or normalize_sheet_name(cat) in l]
            df[LOG_COL] = '정상' if not cat_logs else '\n'.join(cat_logs)
            df.to_excel(writer, sheet_name=safe_name, index=False)

        wb_out = writer.book
        for ws in wb_out.worksheets:
            format_numbers_as_text(ws)
            auto_adjust_column_width(ws)

    print(f'\n저장 완료: {OUTPUT_PATH}')
    print(f'시트 목록: 전체 + {list(merged_sheets.keys())}')
    if logs:
        print(f'\n=== 로그 ({len(logs)}건) ===')
        for l in logs:
            print(f'  {l}')
    else:
        print('\n로그: 정상 (에러 없음)')


if __name__ == '__main__':
    main()

"""
폴더 내 동일 포맷 엑셀 파일들에서 '현금및현금성자산' 시트의 계정과목을 추출해
하나의 엑셀 파일로 합치는 스크립트.

헤더 행: B13~B17 중 '계정과목명' 문자열이 있는 셀이 속한 행.
데이터: 헤더 바로 아래부터 빈 행이 나오기 전까지.

사용법:
    python merge_accounts.py [입력폴더] [출력파일]

    인자 생략 시 현재 폴더의 .xlsx 파일들을 읽어 merged_accounts.xlsx 생성.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAME = "현금및현금성자산"
HEADER_KEYWORD = "계정과목명"
HEADER_SEARCH_ROWS = range(13, 18)  # B13 ~ B17
HEADER_COL = "B"

# 회사명 추출 설정
LEADING_STRIP_PATTERN = re.compile(r"^[0-9.#\-_]+")
TOKEN_SPLIT_PATTERN = re.compile(r"[_\s]+")
# 부분 일치(substring)로 제거할 패턴 — 토큰 안에 포함되면 토큰 전체 제거
DROP_SUBSTRINGS = ("template", "템플릿", "결산자료요청", "공정가치반영")
# 부분 매칭 정규식 — 토큰 안에서 매칭되면 토큰 전체 제거
DROP_CONTAINS_REGEXES = (
    re.compile(r"v[\d.]+", re.IGNORECASE),  # 버전 표기 v1, v2.0 등
    re.compile(r"\d{6,}"),                   # 6자 이상 연속 숫자 (날짜 형식 포함)
)
# 전체 일치 정규식 — 토큰 전체가 매칭될 때만 제거
DROP_FULLMATCH_REGEXES = (
    re.compile(r"[A-Za-z]{3}"),  # 영어 3글자만으로 이루어진 토큰
)
# 정확히 일치할 때만 제거할 토큰
DROP_EXACT = ("CB",)


def _should_drop_token(token: str) -> bool:
    if not token or token.isdigit():
        return True
    if token in DROP_EXACT:
        return True
    if any(p.fullmatch(token) for p in DROP_FULLMATCH_REGEXES):
        return True
    lower = token.lower()
    if any(s in lower for s in DROP_SUBSTRINGS):
        return True
    return any(p.search(token) for p in DROP_CONTAINS_REGEXES)


def extract_company_name(filename: str) -> str:
    """파일명에서 회사명 추출.

    규칙:
      1) 확장자 제거 후 앞쪽 [0-9.#-_] 연속 문자 제거
      2) '_' 또는 공백으로 split
      3) 토큰 제거 규칙:
         - 숫자로만 이루어진 토큰
         - 'template'/'템플릿'/'결산자료요청'/'공정가치반영'이 포함된 토큰
         - 'v1', 'v2.0' 등 버전 표기 포함 토큰
         - 6자 이상 연속 숫자를 포함한 토큰 (260307xx 같은 날짜형)
         - 영어 3글자만으로 이루어진 토큰
         - 정확히 'CB'인 토큰
      4) 남은 토큰을 '_'로 이어 반환
    """
    stem = Path(filename).stem
    stripped = LEADING_STRIP_PATTERN.sub("", stem)
    tokens = [t for t in TOKEN_SPLIT_PATTERN.split(stripped) if t]
    kept = [t for t in tokens if not _should_drop_token(t)]
    return "_".join(kept) if kept else stem


# 출력 스타일
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
MAX_COL_WIDTH = 60


def _cell_display_width(value) -> int:
    """한글/전각문자는 2칸으로 계산."""
    if value is None:
        return 0
    return sum(2 if ord(c) > 127 else 1 for c in str(value))


def format_worksheet(ws) -> None:
    """헤더 스타일 지정 + 열 너비 자동 조정."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_width = 0
        for row_idx in range(1, ws.max_row + 1):
            w = _cell_display_width(ws.cell(row=row_idx, column=col_idx).value)
            if w > max_width:
                max_width = w
        ws.column_dimensions[letter].width = min(max_width + 2, MAX_COL_WIDTH)


def find_header_row(ws) -> int | None:
    """B13~B17 중 '계정과목명'이 있는 행 번호 반환. 못 찾으면 None."""
    for row in HEADER_SEARCH_ROWS:
        value = ws[f"{HEADER_COL}{row}"].value
        if value is not None and HEADER_KEYWORD in str(value):
            return row
    return None


def extract_accounts(xlsx_path: Path) -> pd.DataFrame:
    """엑셀 파일 하나에서 계정과목 표를 DataFrame으로 반환."""
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"시트 '{SHEET_NAME}'를 찾을 수 없습니다.")
        ws = wb[SHEET_NAME]

        header_row = find_header_row(ws)
        if header_row is None:
            raise ValueError(
                f"B13~B17에서 '{HEADER_KEYWORD}' 헤더를 찾지 못했습니다."
            )

        # 헤더 행 전체 값 수집 (B열부터 오른쪽 끝까지)
        headers: list[str] = []
        last_col = 0
        for col_idx, cell in enumerate(ws[header_row], start=1):
            val = cell.value
            if col_idx < 2:  # A열은 무시
                continue
            if val is None or str(val).strip() == "":
                if not headers:
                    continue
                break
            headers.append(str(val).strip())
            last_col = col_idx

        if not headers:
            raise ValueError("헤더 값이 비어있습니다.")

        # 데이터 수집 (헤더 바로 아래부터 빈 행까지)
        rows: list[list] = []
        for row_cells in ws.iter_rows(
            min_row=header_row + 1, min_col=2, max_col=last_col
        ):
            values = [c.value for c in row_cells]
            if all(v is None or str(v).strip() == "" for v in values):
                break
            rows.append(values)

        df = pd.DataFrame(rows, columns=headers)
        df.insert(0, "회사명", extract_company_name(xlsx_path.name))
        df.insert(1, "파일명", xlsx_path.name)
        return df
    finally:
        wb.close()


def merge_folder(input_dir: Path, output_path: Path) -> None:
    xlsx_files = sorted(
        p for p in input_dir.glob("*.xlsx")
        if not p.name.startswith("~$") and p.resolve() != output_path.resolve()
    )
    if not xlsx_files:
        print(f"[경고] '{input_dir}' 폴더에 .xlsx 파일이 없습니다.")
        return

    frames: list[pd.DataFrame] = []
    for path in xlsx_files:
        try:
            df = extract_accounts(path)
            print(f"[OK] {path.name}: {len(df)}행 추출")
            frames.append(df)
        except Exception as e:
            print(f"[실패] {path.name}: {e}")

    if not frames:
        print("추출된 데이터가 없어 출력 파일을 생성하지 않습니다.")
        return

    merged = pd.concat(frames, ignore_index=True)
    sheet = "계정과목_통합"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name=sheet, index=False)
        format_worksheet(writer.sheets[sheet])
    print(f"\n[완료] {output_path} ({len(merged)}행)")


def main() -> None:
    input_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
    output_path = (
        Path(sys.argv[2]) if len(sys.argv) > 2 else input_dir / "merged_accounts.xlsx"
    )

    if not input_dir.is_dir():
        print(f"[오류] 입력 폴더가 존재하지 않습니다: {input_dir}")
        sys.exit(1)

    merge_folder(input_dir, output_path)


if __name__ == "__main__":
    main()

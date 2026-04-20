"""
폴더 내 동일 포맷 엑셀 파일들에서 지정 시트의 계정과목을 추출해
하나의 엑셀 파일로 시트별로 합치는 스크립트.

대상 시트: 현금및현금성자산(계정과목명), 유동화자산(계정과목코드명),
           장단기및유동화부채(계정과목명)
헤더 행: 각 시트의 B10~B40 중 해당 키워드가 포함된 셀이 있는 행.
데이터: 헤더 바로 아래부터 빈 행이 나오기 전까지.

사용법:
    python merge_spc.py [입력폴더] [출력파일]

    인자 생략 시 스크립트와 같은 위치의 SPC 폴더를 읽고,
    스크립트와 같은 위치에 merge_spc.xlsx를 생성한다.
"""

from __future__ import annotations

import re
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 시트명 -> 헤더 탐색 키워드
SHEET_HEADER_KEYWORDS = {
    "현금및현금성자산": "계정과목명",
    "유동화자산": "계정과목코드명",
    "장단기및유동화부채": "계정과목명",
}
SHEET_NAMES = tuple(SHEET_HEADER_KEYWORDS.keys())
HEADER_SEARCH_ROWS = range(10, 41)  # B10 ~ B40
HEADER_COL = "B"

# FALSE 값 점검 범위 (전체 시트의 E1:F40)
FALSE_CHECK_SHEET = "FALSE check"
FALSE_CHECK_MIN_COL = 5  # E
FALSE_CHECK_MAX_COL = 6  # F
FALSE_CHECK_MAX_ROW = 40

ERROR_SHEET = "ERROR"

# 회사명 추출 설정
LEADING_STRIP_PATTERN = re.compile(r"^[0-9.#\-_]+")
TOKEN_SPLIT_PATTERN = re.compile(r"[_\s]+")
# 부분 일치(substring)로 제거할 패턴 — 토큰 안에 포함되면 토큰 전체 제거
DROP_SUBSTRINGS = ("template", "템플릿", "결산자료요청", "공정가치반영")
# 부분 매칭 정규식 — 토큰 안에서 매칭되면 토큰 전체 제거
DROP_CONTAINS_REGEXES = (
    re.compile(r"v[\d.]+", re.IGNORECASE),  # 버전 표기 v1, v2.0 등
    re.compile(r"\d{6,}"),                   # 6자 이상 연속 숫자 (날짜 형식 포함)
)
# 전체 일치 정규식 — 토큰 전체가 매칭될 때만 제거
DROP_FULLMATCH_REGEXES = (
    re.compile(r"[A-Za-z]{3}"),  # 영어 3글자만으로 이루어진 토큰
)
# 정확히 일치할 때만 제거할 토큰
DROP_EXACT = ("CB",)


def _should_drop_token(token: str) -> bool:
    if not token or token.isdigit():
        return True
    if token in DROP_EXACT:
        return True
    if any(p.fullmatch(token) for p in DROP_FULLMATCH_REGEXES):
        return True
    lower = token.lower()
    if any(s in lower for s in DROP_SUBSTRINGS):
        return True
    return any(p.search(token) for p in DROP_CONTAINS_REGEXES)


def extract_company_name(filename: str) -> str:
    """파일명에서 회사명 추출.

    규칙:
      1) 확장자 제거 후 앞쪽 [0-9.#-_] 연속 문자 제거
      2) '_' 또는 공백으로 split
      3) 토큰 제거 규칙:
         - 숫자로만 이루어진 토큰
         - 'template'/'템플릿'/'결산자료요청'/'공정가치반영'이 포함된 토큰
         - 'v1', 'v2.0' 등 버전 표기 포함 토큰
         - 6자 이상 연속 숫자를 포함한 토큰 (260307xx 같은 날짜형)
         - 영어 3글자만으로 이루어진 토큰
         - 정확히 'CB'인 토큰
      4) 남은 토큰을 '_'로 이어 반환
    """
    stem = Path(filename).stem
    stripped = LEADING_STRIP_PATTERN.sub("", stem)
    tokens = [t for t in TOKEN_SPLIT_PATTERN.split(stripped) if t]
    kept = [t for t in tokens if not _should_drop_token(t)]
    return "_".join(kept) if kept else stem


# 출력 스타일
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_SIDE = Side(style="thin", color="000000")
CELL_BORDER = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
# 헤더 키워드가 컬럼명에 포함되면 지정 셀서식 적용 (위에서부터 첫 매칭)
COLUMN_FORMAT_RULES: tuple[tuple[tuple[str, ...], str], ...] = (
    (("금액", "잔액", "수익"), "#,##0;[Red](#,##0)"),  # 숫자-빨간색괄호
    (("실행일", "만기", "발행일"), "yyyy-mm-dd"),        # 간단한 날짜 (만기·만기일·만기일자 등)
    (("이자율",), "0.00%"),                             # 백분율 (소숫점 둘째자리)
)
# 빈 값을 0으로 채울 컬럼 키워드
ZERO_FILL_KEYWORDS = ("금액", "잔액", "수익")

# 특정 셀서식이 적용된 컬럼의 최소 너비 보장 (발행일자/만기일자 등 좁음 방지)
FORMAT_MIN_WIDTHS = {
    "yyyy-mm-dd": 14,
    "0.00%": 10,
}
MAX_COL_WIDTH = 60


def _cell_display_width(value) -> int:
    """한글/전각문자는 2칸으로 계산."""
    if value is None:
        return 0
    return sum(2 if ord(c) > 127 else 1 for c in str(value))


def _resolve_column_format(header: str) -> str | None:
    """헤더 문자열에 맞는 셀서식 반환 (매칭 없으면 None)."""
    for keywords, fmt in COLUMN_FORMAT_RULES:
        if any(kw in header for kw in keywords):
            return fmt
    return None


def format_worksheet(ws) -> None:
    """헤더 스타일 + 전체 테두리 + 헤더별 셀서식 + 열 너비 자동 조정."""
    # 컬럼 인덱스별 적용할 셀서식 (없으면 None)
    col_formats: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=c).value or "")
        fmt = _resolve_column_format(header)
        if fmt:
            col_formats[c] = fmt

    # 헤더 스타일
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # 전체 셀 테두리 + 헤더별 셀서식
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = CELL_BORDER
            if row_idx > 1 and col_idx in col_formats:
                cell.number_format = col_formats[col_idx]

    # 열 너비 자동 조정
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_width = 0
        for row_idx in range(1, ws.max_row + 1):
            w = _cell_display_width(ws.cell(row=row_idx, column=col_idx).value)
            if w > max_width:
                max_width = w
        width = min(max_width + 2, MAX_COL_WIDTH)
        fmt = col_formats.get(col_idx)
        if fmt in FORMAT_MIN_WIDTHS:
            width = max(width, FORMAT_MIN_WIDTHS[fmt])
        ws.column_dimensions[letter].width = width


def find_header_row(ws, keyword: str) -> int | None:
    """B10~B40 중 주어진 키워드가 있는 행 번호 반환. 못 찾으면 None."""
    for row in HEADER_SEARCH_ROWS:
        value = ws[f"{HEADER_COL}{row}"].value
        if value is not None and keyword in str(value):
            return row
    return None


def _is_false_value(v) -> bool:
    if v is False:
        return True
    if isinstance(v, str) and v.strip().upper() == "FALSE":
        return True
    return False


def collect_false_cells(wb, xlsx_path: Path) -> list[dict]:
    """전체 시트의 E1:F40에서 FALSE 값을 가진 셀 위치 수집."""
    company = extract_company_name(xlsx_path.name)
    results: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row_cells in ws.iter_rows(
            min_row=1, max_row=FALSE_CHECK_MAX_ROW,
            min_col=FALSE_CHECK_MIN_COL, max_col=FALSE_CHECK_MAX_COL,
        ):
            for cell in row_cells:
                if _is_false_value(cell.value):
                    results.append({
                        "회사명": company,
                        "파일명": xlsx_path.name,
                        "시트": sheet,
                        "셀위치": cell.coordinate,
                    })
    return results


def extract_sheet(wb, sheet_name: str, xlsx_path: Path) -> pd.DataFrame:
    """특정 시트에서 계정과목 표를 DataFrame으로 반환."""
    ws = wb[sheet_name]
    keyword = SHEET_HEADER_KEYWORDS[sheet_name]

    header_row = find_header_row(ws, keyword)
    if header_row is None:
        raise ValueError(
            f"B10~B40에서 '{keyword}' 헤더를 찾지 못했습니다."
        )

    # 헤더 행 전체 값 수집 (B열부터 오른쪽 끝까지)
    headers: list[str] = []
    last_col = 0
    for col_idx, cell in enumerate(ws[header_row], start=1):
        val = cell.value
        if col_idx < 2:  # A열은 무시
            continue
        if val is None or str(val).strip() == "":
            if not headers:
                continue
            break
        headers.append(str(val).strip())
        last_col = col_idx

    if not headers:
        raise ValueError("헤더 값이 비어있습니다.")

    # 데이터 수집 (헤더 바로 아래부터 빈 행까지)
    rows: list[list] = []
    for row_cells in ws.iter_rows(
        min_row=header_row + 1, min_col=2, max_col=last_col
    ):
        values = [c.value for c in row_cells]
        if all(v is None or str(v).strip() == "" for v in values):
            break
        rows.append(values)

    df = pd.DataFrame(rows, columns=headers)

    # 금액/잔액/수익 컬럼은 빈 값(None/NaN/빈 문자열)을 0으로 채움
    for col in df.columns:
        if any(kw in str(col) for kw in ZERO_FILL_KEYWORDS):
            df[col] = df[col].replace({"": None}).fillna(0)

    df.insert(0, "회사명", extract_company_name(xlsx_path.name))
    df.insert(1, "파일명", xlsx_path.name)
    return df


def merge_folder(input_dir: Path, output_path: Path) -> None:
    xlsx_files = sorted(
        p for p in input_dir.glob("*.xlsx")
        if not p.name.startswith("~$") and p.resolve() != output_path.resolve()
    )
    if not xlsx_files:
        print(f"[경고] '{input_dir}' 폴더에 .xlsx 파일이 없습니다.")
        return

    # sheet_name -> list[DataFrame]
    sheet_frames: dict[str, list[pd.DataFrame]] = {s: [] for s in SHEET_NAMES}
    false_rows: list[dict] = []
    error_rows: list[dict] = []

    def _record_warnings(caught, company, filename, sheet, phase):
        for w in caught:
            msg = f"{type(w.message).__name__}: {w.message}"
            if phase:
                msg = f"[{phase}] {msg}"
            print(f"[경고] {filename} / {sheet or '-'}: {msg}")
            error_rows.append({
                "회사명": company, "파일명": filename,
                "시트": sheet, "사유": msg,
            })

    total = len(xlsx_files)
    for idx, path in enumerate(xlsx_files, start=1):
        prog = f"{idx}/{total}"
        company = extract_company_name(path.name)

        # 파일 열기 (경고 포함)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            try:
                wb = load_workbook(filename=path, data_only=True, read_only=True)
                open_error = None
            except Exception as e:
                wb = None
                open_error = e
        _record_warnings(caught, company, path.name, "", "파일 열기")
        if open_error is not None:
            print(f"[실패] {path.name} 열기 오류: {open_error}")
            error_rows.append({
                "회사명": company, "파일명": path.name,
                "시트": "", "사유": f"파일 열기 실패: {open_error}",
            })
            continue

        try:
            for sheet in SHEET_NAMES:
                if sheet not in wb.sheetnames:
                    print(f"[건너뜀] {path.name} / {sheet}: 시트 없음")
                    error_rows.append({
                        "회사명": company, "파일명": path.name,
                        "시트": sheet, "사유": "시트 없음",
                    })
                    continue
                with warnings.catch_warnings(record=True) as caught:
                    warnings.simplefilter("always")
                    try:
                        df = extract_sheet(wb, sheet, path)
                        print(f"[OK {prog}] {path.name} / {sheet}: {len(df)}행 추출")
                        sheet_frames[sheet].append(df)
                    except Exception as e:
                        print(f"[실패] {path.name} / {sheet}: {e}")
                        error_rows.append({
                            "회사명": company, "파일명": path.name,
                            "시트": sheet, "사유": str(e),
                        })
                _record_warnings(caught, company, path.name, sheet, "시트 처리")

            # FALSE 점검 (경고 포함)
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                hits = collect_false_cells(wb, path)
            _record_warnings(caught, company, path.name, "(전체)", "FALSE 점검")
            if hits:
                print(f"[FALSE {prog}] {path.name}: {len(hits)}건")
                false_rows.extend(hits)
        finally:
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                wb.close()
            _record_warnings(caught, company, path.name, "", "파일 닫기")

    non_empty = {s: fs for s, fs in sheet_frames.items() if fs}
    if not non_empty and not false_rows and not error_rows:
        print("추출된 데이터가 없어 출력 파일을 생성하지 않습니다.")
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet, frames in non_empty.items():
            merged = pd.concat(frames, ignore_index=True)
            merged.to_excel(writer, sheet_name=sheet, index=False)
            format_worksheet(writer.sheets[sheet])
            print(f"  → {sheet}: {len(merged)}행")

        false_df = pd.DataFrame(
            false_rows, columns=["회사명", "파일명", "시트", "셀위치"]
        )
        false_df.to_excel(writer, sheet_name=FALSE_CHECK_SHEET, index=False)
        format_worksheet(writer.sheets[FALSE_CHECK_SHEET])
        print(f"  → {FALSE_CHECK_SHEET}: {len(false_df)}건")

        error_df = pd.DataFrame(
            error_rows, columns=["회사명", "파일명", "시트", "사유"]
        )
        error_df.to_excel(writer, sheet_name=ERROR_SHEET, index=False)
        format_worksheet(writer.sheets[ERROR_SHEET])
        print(f"  → {ERROR_SHEET}: {len(error_df)}건")
    print(f"\n[완료] {output_path}")


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = SCRIPT_DIR / "SPC"
DEFAULT_OUTPUT_PATH = SCRIPT_DIR / "merge_spc.xlsx"


def main() -> None:
    input_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT_DIR
    output_path = (
        Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT_PATH
    )

    if not input_dir.is_dir():
        print(f"[오류] 입력 폴더가 존재하지 않습니다: {input_dir}")
        sys.exit(1)

    merge_folder(input_dir, output_path)


if __name__ == "__main__":
    main()
